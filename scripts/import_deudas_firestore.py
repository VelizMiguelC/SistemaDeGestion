#!/usr/bin/env python3
"""
Importador masivo: 'DEUDAS Y DEUDORES (1).xlsx' -> Cloud Firestore
===================================================================

Lee las hojas 'Clientes', 'Deudas' y 'Pagos' del Excel y las sube a
Firestore como tres colecciones de nivel superior:

    clientes/{ID_Cliente}  -> { nombre, telefono, notas }
    deudas/{ID_Deuda}      -> { idCliente, concepto, montoTotal, fechaEmision, estado }
    pagos/{ID_Pago}        -> { idDeuda, fechaPago, montoAbonado, metodoPago }

Requisitos
----------
    pip install firebase-admin openpyxl

Credenciales
------------
Este script usa el SDK de administrador (firebase-admin), que necesita una
clave de cuenta de servicio (Service Account) con permisos sobre Firestore
del proyecto "inventario-app-abe77":

    1. Andá a Firebase Console -> ⚙️ Configuración del proyecto
       -> Cuentas de servicio -> "Generar nueva clave privada".
    2. Descargá el archivo .json y guardalo, por ejemplo, como
       `serviceAccountKey.json` en la raíz del proyecto Flutter
       (¡NO lo subas a git! Ya está en .gitignore si usás el de Flutter,
       pero conviene confirmarlo).
    3. Corré el script apuntando a ese archivo (ver ejemplos más abajo).

⚠️  Por seguridad, esa clave nunca debe pegarse en un chat ni commitearse:
    da acceso total de administrador a tu proyecto de Firebase/GCP.

Uso
---
    # 1) Vista previa SIN tocar Firestore (recomendado primero):
    python3 import_deudas_firestore.py --dry-run

    # 2) Importación real:
    python3 import_deudas_firestore.py \
        --service-account serviceAccountKey.json \
        --excel "DEUDAS Y DEUDORES (1).xlsx"

Por defecto busca el Excel como "DEUDAS Y DEUDORES (1).xlsx" y la clave
como "serviceAccountKey.json", ambos en el directorio actual.
"""

from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError:
    sys.exit(
        "Falta la librería 'openpyxl'. Instalala con:\n"
        "    pip install openpyxl"
    )

# Tamaño máximo de un batch write en Firestore (el límite real es 500;
# dejamos margen).
BATCH_SIZE = 400


# ---------------------------------------------------------------------------
# Utilidades de limpieza de datos
# ---------------------------------------------------------------------------

def limpiar_texto(valor: Any) -> str | None:
    """Convierte a string, recorta espacios y descarta vacíos -> None."""
    if valor is None:
        return None
    texto = str(valor).strip()
    return texto or None


def limpiar_telefono(valor: Any) -> str | None:
    """Normaliza un teléfono a texto, sin perder ceros a la izquierda.

    Excel a veces guarda los teléfonos como número (float/int) en vez de
    texto; en ese caso hay que evitar que quede como "3816636750.0".
    """
    if valor is None:
        return None
    if isinstance(valor, float):
        if valor.is_integer():
            valor = int(valor)
    texto = str(valor).strip()
    if texto.endswith(".0"):
        texto = texto[:-2]
    return texto or None


def limpiar_monto(valor: Any) -> float:
    """Convierte a float, tratando vacíos/None como 0."""
    if valor is None or valor == "":
        return 0.0
    try:
        return float(valor)
    except (TypeError, ValueError):
        return 0.0


def a_datetime_utc(valor: Any) -> datetime | None:
    """Normaliza fechas leídas de Excel a datetime aware en UTC."""
    if valor is None:
        return None
    if isinstance(valor, datetime):
        dt = valor
    else:
        # Por si la celda vino como texto ("2026-07-25") en vez de fecha.
        texto = str(valor).strip()
        if not texto:
            return None
        try:
            dt = datetime.fromisoformat(texto)
        except ValueError:
            return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt


def quitar_nulos(doc: dict[str, Any]) -> dict[str, Any]:
    """Elimina claves con valor None para no ensuciar Firestore con nulls."""
    return {k: v for k, v in doc.items() if v is not None}


# ---------------------------------------------------------------------------
# Lectura del Excel
# ---------------------------------------------------------------------------

@dataclass
class ResultadoImportacion:
    clientes: dict[str, dict[str, Any]] = field(default_factory=dict)
    deudas: dict[str, dict[str, Any]] = field(default_factory=dict)
    pagos: dict[str, dict[str, Any]] = field(default_factory=dict)
    avisos: list[str] = field(default_factory=list)


def leer_hoja(ws) -> list[dict[str, Any]]:
    """Devuelve las filas de una hoja como lista de dicts {columna: valor},
    usando la primera fila como encabezado e ignorando filas vacías."""
    filas = list(ws.iter_rows(values_only=True))
    if not filas:
        return []
    encabezado = [str(c).strip() if c is not None else "" for c in filas[0]]
    resultado = []
    for fila in filas[1:]:
        if all(v is None or str(v).strip() == "" for v in fila):
            continue
        resultado.append(dict(zip(encabezado, fila)))
    return resultado


def procesar_excel(ruta_excel: Path) -> ResultadoImportacion:
    wb = openpyxl.load_workbook(ruta_excel, data_only=True)
    resultado = ResultadoImportacion()

    for hoja_requerida in ("Clientes", "Deudas", "Pagos"):
        if hoja_requerida not in wb.sheetnames:
            sys.exit(
                f"El Excel no tiene una hoja llamada '{hoja_requerida}'. "
                f"Hojas encontradas: {wb.sheetnames}"
            )

    # --- Clientes ------------------------------------------------------
    for fila in leer_hoja(wb["Clientes"]):
        id_cliente = limpiar_texto(fila.get("ID_Cliente"))
        if not id_cliente:
            resultado.avisos.append(
                f"[Clientes] Fila sin ID_Cliente, se omite: {fila}"
            )
            continue
        nombre = limpiar_texto(fila.get("Nombre_Cliente"))
        if not nombre:
            resultado.avisos.append(
                f"[Clientes] {id_cliente} no tiene Nombre_Cliente; se guarda vacío."
            )
        doc = quitar_nulos({
            "nombre": nombre or "",
            "telefono": limpiar_telefono(fila.get("Telefono")),
            "notas": limpiar_texto(fila.get("Notas")),
        })
        resultado.clientes[id_cliente] = doc

    # --- Deudas ----------------------------------------------------------
    for fila in leer_hoja(wb["Deudas"]):
        id_deuda = limpiar_texto(fila.get("ID_Deuda"))
        if not id_deuda:
            resultado.avisos.append(f"[Deudas] Fila sin ID_Deuda, se omite: {fila}")
            continue

        id_cliente = limpiar_texto(fila.get("ID_Cliente"))
        if id_cliente and id_cliente not in resultado.clientes:
            resultado.avisos.append(
                f"[Deudas] {id_deuda} referencia ID_Cliente '{id_cliente}' "
                "que no existe en la hoja Clientes (se importa igual)."
            )

        fecha_emision = a_datetime_utc(fila.get("Fecha_Emision"))
        estado = limpiar_texto(fila.get("Estado")) or "Pendiente"

        doc = quitar_nulos({
            "idCliente": id_cliente,
            "concepto": limpiar_texto(fila.get("Concepto_Detalle")) or "",
            "montoTotal": limpiar_monto(fila.get("Monto_Total")),
            "fechaEmision": fecha_emision,
            "estado": estado,
        })
        resultado.deudas[id_deuda] = doc

    # --- Pagos -------------------------------------------------------------
    for fila in leer_hoja(wb["Pagos"]):
        id_pago = limpiar_texto(fila.get("ID_Pago"))
        if not id_pago:
            resultado.avisos.append(f"[Pagos] Fila sin ID_Pago, se omite: {fila}")
            continue

        id_deuda = limpiar_texto(fila.get("ID_Deuda"))
        if id_deuda and id_deuda not in resultado.deudas:
            resultado.avisos.append(
                f"[Pagos] {id_pago} referencia ID_Deuda '{id_deuda}' que no "
                "existe en la hoja Deudas (se importa igual)."
            )

        doc = quitar_nulos({
            "idDeuda": id_deuda,
            "fechaPago": a_datetime_utc(fila.get("Fecha_Pago")),
            "montoAbonado": limpiar_monto(fila.get("Monto_Abonado")),
            "metodoPago": limpiar_texto(fila.get("Metodo_Pago")) or "Efectivo",
        })
        resultado.pagos[id_pago] = doc

    return resultado


# ---------------------------------------------------------------------------
# Subida a Firestore
# ---------------------------------------------------------------------------

def subir_a_firestore(resultado: ResultadoImportacion, service_account_path: Path) -> None:
    try:
        import firebase_admin
        from firebase_admin import credentials, firestore
    except ImportError:
        sys.exit(
            "Falta la librería 'firebase-admin'. Instalala con:\n"
            "    pip install firebase-admin"
        )

    if not service_account_path.exists():
        sys.exit(
            f"No se encontró la clave de cuenta de servicio en "
            f"'{service_account_path}'. Ver instrucciones al inicio de este "
            "archivo para generarla desde Firebase Console."
        )

    cred = credentials.Certificate(str(service_account_path))
    if not firebase_admin._apps:
        firebase_admin.initialize_app(cred)
    db = firestore.client()

    def subir_coleccion(nombre_coleccion: str, docs: dict[str, dict[str, Any]]) -> None:
        items = list(docs.items())
        total = len(items)
        subidos = 0
        for inicio in range(0, total, BATCH_SIZE):
            lote = items[inicio:inicio + BATCH_SIZE]
            batch = db.batch()
            for doc_id, data in lote:
                ref = db.collection(nombre_coleccion).document(doc_id)
                batch.set(ref, data, merge=True)
            batch.commit()
            subidos += len(lote)
            print(f"  {nombre_coleccion}: {subidos}/{total} documentos subidos...")
        print(f"✔ {nombre_coleccion}: {total} documentos importados.")

    print("Subiendo a Firestore...")
    subir_coleccion("clientes", resultado.clientes)
    subir_coleccion("deudas", resultado.deudas)
    subir_coleccion("pagos", resultado.pagos)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def mostrar_resumen(resultado: ResultadoImportacion, mostrar_muestras: bool = True) -> None:
    print("Resumen de datos leídos del Excel")
    print("-" * 40)
    print(f"  clientes: {len(resultado.clientes)}")
    print(f"  deudas:   {len(resultado.deudas)}")
    print(f"  pagos:    {len(resultado.pagos)}")

    if resultado.avisos:
        print()
        print(f"⚠ Avisos ({len(resultado.avisos)}):")
        for aviso in resultado.avisos:
            print(f"  - {aviso}")

    if mostrar_muestras:
        print()
        print("Muestra (primeros 2 documentos por colección):")
        for nombre, docs in (
            ("clientes", resultado.clientes),
            ("deudas", resultado.deudas),
            ("pagos", resultado.pagos),
        ):
            print(f"  {nombre}:")
            for doc_id, data in list(docs.items())[:2]:
                print(f"    {doc_id}: {data}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Importa 'DEUDAS Y DEUDORES.xlsx' a Cloud Firestore.",
    )
    parser.add_argument(
        "--excel",
        default="DEUDAS Y DEUDORES (1).xlsx",
        help="Ruta al archivo Excel (por defecto: %(default)s)",
    )
    parser.add_argument(
        "--service-account",
        default="serviceAccountKey.json",
        help="Ruta a la clave de cuenta de servicio de Firebase (por defecto: %(default)s)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Solo lee y valida el Excel; no escribe nada en Firestore.",
    )
    args = parser.parse_args()

    ruta_excel = Path(args.excel)
    if not ruta_excel.exists():
        sys.exit(f"No se encontró el archivo Excel: {ruta_excel}")

    resultado = procesar_excel(ruta_excel)
    mostrar_resumen(resultado)

    if args.dry_run:
        print()
        print("Modo --dry-run: no se escribió nada en Firestore.")
        return

    print()
    confirmacion = input(
        f"¿Confirmás subir {len(resultado.clientes)} clientes, "
        f"{len(resultado.deudas)} deudas y {len(resultado.pagos)} pagos "
        "a Firestore? [s/N]: "
    ).strip().lower()
    if confirmacion != "s":
        print("Cancelado. No se subió nada.")
        return

    subir_a_firestore(resultado, Path(args.service_account))
    print()
    print("Importación completa.")


if __name__ == "__main__":
    main()
